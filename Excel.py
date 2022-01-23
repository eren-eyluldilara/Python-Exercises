from openpyxl import workbook,load_workbook,worksheet
dosya=load_workbook("/Users/eylul/Desktop/Eylul/Kitap1.xlsx")
dosya.active=0
ws=dosya.active
ws['B1'].value="5"
dosya.save("/Users/eylul/Desktop/Eylul/Kitap1.xlsx")
dosya.close()
