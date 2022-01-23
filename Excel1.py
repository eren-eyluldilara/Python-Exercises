from openpyxl import workbook,load_workbook,worksheet
dosya=load_workbook("/Users/eylul/Desktop/Eylul/Kitap1.xlsx")
dosya.active=0
ws=dosya.active
for i in range(1,6):
 ws.cell(row=i,column =2,value="deneme")
dosya.save("/Users/eylul/Desktop/Eylul/Kitap1.xlsx")
dosya.close()